import requests
import pandas as pd
import time
import os
from openpyxl import Workbook, load_workbook

# API Endpoint for fetching the top 50 cryptocurrencies by market cap
API_URL = "https://api.coingecko.com/api/v3/coins/markets?vs_currency=usd&order=market_cap_desc&per_page=50&page=1&sparkline=false"

# Excel file path
EXCEL_FILE = "crypto_live_data.xlsx"

# Function to fetch live data from CoinGecko API
def fetch_crypto_data():
    response = requests.get(API_URL)
    if response.status_code == 200:
        data = response.json()
        crypto_list = []

        for crypto in data:
            crypto_list.append({
                "Name": crypto["name"],
                "Symbol": crypto["symbol"].upper(),
                "Current Price (USD)": crypto["current_price"],
                "Market Cap (USD)": crypto["market_cap"],
                "24h Trading Volume (USD)": crypto["total_volume"],
                "24h Price Change (%)": crypto["price_change_percentage_24h"],
            })
        return pd.DataFrame(crypto_list)
    else:
        print("Error fetching data from API.")
        return pd.DataFrame()

# Perform data analysis
def analyze_data(df):
    analysis = {}

    # Top 5 cryptocurrencies by market cap
    top_5 = df.sort_values(by="Market Cap (USD)", ascending=False).head(5)[["Name", "Market Cap (USD)"]]
    analysis["Top 5 by Market Cap"] = top_5

    # Average price of the top 50 cryptocurrencies
    analysis["Average Price"] = df["Current Price (USD)"].mean()

    # Highest and lowest 24-hour price change
    highest_change = df.loc[df["24h Price Change (%)"].idxmax()]["Name"]
    lowest_change = df.loc[df["24h Price Change (%)"].idxmin()]["Name"]
    
    analysis["Highest 24h Price Change"] = highest_change
    analysis["Lowest 24h Price Change"] = lowest_change

    return analysis

# Write data and analysis to Excel
def update_excel(df, analysis):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]  # Remove existing sheet to clear old data
    
    sheet = wb.create_sheet("Sheet")

    # Write live crypto data
    for col_num, header in enumerate(df.columns, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(df.values, start=2):
        for col_num, value in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num, value=value)

    # Write analysis data
    analysis_start_row = len(df) + 3
    sheet.cell(row=analysis_start_row, column=1, value="Analysis")
    
    sheet.cell(row=analysis_start_row + 1, column=1, value="Top 5 by Market Cap:")
    for row_num, row_data in enumerate(analysis["Top 5 by Market Cap"].values.tolist(), start=analysis_start_row + 2):
        sheet.cell(row=row_num, column=1, value=row_data[0])  # Name
        sheet.cell(row=row_num, column=2, value=row_data[1])  # Market Cap

    sheet.cell(row=analysis_start_row + 8, column=1, value=f"Average Price: ${analysis['Average Price']:.2f}")
    sheet.cell(row=analysis_start_row + 9, column=1, value=f"Highest 24h Price Change: {analysis['Highest 24h Price Change']}")
    sheet.cell(row=analysis_start_row + 10, column=1, value=f"Lowest 24h Price Change: {analysis['Lowest 24h Price Change']}")

    wb.save(EXCEL_FILE)
    wb.close()

# Main function for fetching, analyzing, and updating Excel
def main():
    print("Starting live cryptocurrency data update...")
    while True:
        # Fetch live data
        crypto_data = fetch_crypto_data()
        
        if not crypto_data.empty:
            # Perform analysis
            analysis = analyze_data(crypto_data)

            # Update Excel sheet
            update_excel(crypto_data, analysis)
            print("Excel updated with latest data.")

        # Wait for 5 minutes before the next update
        time.sleep(300)

if __name__ == "__main__":
    main()
